import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os

class  ExcelHistory:
    def __init__(self):
        self.file_name = "history.xlsx"

    def create_or_load_workbook(self):
        if os.path.exists(self.file_name):
            self.workbook = openpyxl.load_workbook(self.file_name)
        else:
            self.workbook = Workbook()
            self.workbook.active.append(["用户名", "问题", "答案", "间距时间（s）"])
            self.save_workbook()

    def save_workbook(self):
        self.workbook.save(self.file_name)

    def insert_data(self, username, question, answer, query_time):
        self.create_or_load_workbook()
        sheet = self.workbook.active
        sheet.append([username, question, answer, query_time])
        self.save_workbook()


