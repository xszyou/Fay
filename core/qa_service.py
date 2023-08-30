
from openpyxl import load_workbook
import difflib
from utils import config_util as cfg
from scheduler.thread_manager import MyThread
import shlex
import subprocess
import time

def question(query_type,text):
    qa = QAService()
    answer = qa.question(query_type,text)
    return answer

def run_script(command):
    args = shlex.split(command)  # 分割命令行参数
    subprocess.Popen(args)

class QAService:
    
    def __init__(self):
         # 人设提问关键字
        self.attribute_keyword = [
            [['你叫什么名字', '你的名字是什么'], 'name'],
            [['你是男的还是女的', '你是男生还是女生', '你的性别是什么', '你是男生吗', '你是女生吗', '你是男的吗', '你是女的吗', '你是男孩子吗', '你是女孩子吗', ], 'gender', ],
            [['你今年多大了', '你多大了', '你今年多少岁', '你几岁了', '你今年几岁了', '你今年几岁了', '你什么时候出生', '你的生日是什么', '你的年龄'], 'age', ],
            [['你的家乡在哪', '你的家乡是什么', '你家在哪', '你住在哪', '你出生在哪', '你的出生地在哪', '你的出生地是什么', ], 'birth', ],
            [['你的生肖是什么', '你属什么', ], 'zodiac', ],
            [['你是什么座', '你是什么星座', '你的星座是什么', ], 'constellation', ],
            [['你是做什么的', '你的职业是什么', '你是干什么的', '你的职位是什么', '你的工作是什么', '你是做什么工作的'], 'job', ],
            [['你的爱好是什么', '你有爱好吗', '你喜欢什么', '你喜欢做什么'], 'hobby'],
            [['联系方式', '联系你们', '怎么联系客服', '有没有客服'], 'contact']
        ]

        self.command_keyword = [
            [['关闭', '再见', '你走吧'], 'stop'],
            [['静音', '闭嘴', '我想静静'], 'mute'],
            [['取消静音', '你在哪呢', '你可以说话了'], 'unmute'],
            [['换个性别', '换个声音'], 'changeVoice']
        ]

    def question(self, query_type, text):
        if query_type == 'qa':
            answer_dict = self.__read_qna(cfg.config['interact']['QnA'])
            answer, action = self.__get_keyword(answer_dict, text, query_type)
            if action:
                MyThread(target=self.__run, args=[action]).start()
            return answer
    
        elif query_type == 'Persona':
            answer_dict = self.attribute_keyword
            answer, action  = self.__get_keyword(answer_dict, text, query_type)
        elif query_type == 'command':
            answer, action  = self.__get_keyword(self.command_keyword, text, query_type)
        return answer

    def __run(self,action):
        time.sleep(2)
        run_script(action)   

    def __read_qna(self, filename):
        qna = []
        try:
            wb = load_workbook(filename)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= 2:
                    qna.append([row[0].split(";"), row[1], row[2] if len(row) >= 3 else None])
        except BaseException as e:
            print(f"无法读取Q&A文件 {filename} -> {e}")
        return qna

    def __get_keyword(self, keyword_dict, text, query_type):
        last_similar = 0
        last_answer = ''
        last_action = ''
        for qa in keyword_dict:
            for quest in qa[0]:
                similar = self.__string_similar(text, quest)
                if quest in text:
                    similar += 0.3
                if similar > last_similar:
                    last_similar = similar
                    last_answer = qa[1]
                    if query_type == "qa":
                        last_action = qa[2]
        if last_similar >= 0.6:
            return last_answer, last_action
        return None, None

    def __string_similar(self, s1, s2):
        return difflib.SequenceMatcher(None, s1, s2).quick_ratio()


